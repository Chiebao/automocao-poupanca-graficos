import os
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
import calendar

# ---------- Configurações ----------
ARQUIVO_EXCEL = "valores_diarios.xlsx"
NOME_ABA = "Economia"
COR_FUNDO = "#252525"
COR_TEXTO = "#FFFFFF"
COR_DESTAQUE = "#03fcdf"

# ---------- Abrir ou criar planilha ----------
if os.path.exists(ARQUIVO_EXCEL):
    wb = load_workbook(ARQUIVO_EXCEL)
else:
    wb = Workbook()

if NOME_ABA in wb.sheetnames:
    sheet = wb[NOME_ABA]
else:
    sheet = wb.active
    sheet.title = NOME_ABA

# Cabeçalhos
if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
    sheet.cell(row=1, column=1, value="Data")
    sheet.cell(row=1, column=2, value="Valor")

# Valores já salvos
valores_salvos = [float(c.value) for c in sheet["B"][1:] if c.value is not None]

# ---------- Funções ----------
def atualizar_total():
    total = sum(valores_salvos)
    label_total.config(text=f"Total economizado: R${total:.2f}")

def salvar_valor():
    try:
        valor = float(entry_valor.get().replace(",", "."))
    except ValueError:
        label_status.config(text="Digite um valor numérico válido!", fg="red", bg=COR_FUNDO)
        return

    valores_salvos.append(valor)

    proxima_linha = sheet.max_row + 1
    sheet.cell(row=proxima_linha, column=1, value=date.today().strftime("%d-%m-%Y"))
    sheet.cell(row=proxima_linha, column=2, value=valor)
    wb.save(ARQUIVO_EXCEL)

    entry_valor.delete(0, tk.END)
    label_status.config(text="Valor salvo com sucesso :)", fg=COR_DESTAQUE, bg=COR_FUNDO)
    atualizar_total()
    atualizar_graficos()  # atualiza gráficos automaticamente

def atualizar_graficos():
    """Atualiza os gráficos no frame_grafico sem esconder o restante da interface."""
    # Limpa gráfico antigo
    for widget in frame_grafico.winfo_children():
        widget.destroy()

    # Carrega dados do Excel
    wb_local = load_workbook(ARQUIVO_EXCEL)
    ws = wb_local[NOME_ABA]

    datas, valores = [], []
    for data_cell, valor_cell in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if data_cell is None or valor_cell is None:
            continue
        if isinstance(data_cell, datetime):
            data = data_cell.date()
        elif isinstance(data_cell, str):
            try:
                data = datetime.strptime(data_cell, "%d-%m-%Y").date()
            except ValueError:
                continue
        elif isinstance(data_cell, (int, float)):
            data = from_excel(data_cell).date()
        else:
            continue
        datas.append(data)
        valores.append(float(valor_cell))

    if not valores:
        aviso = tk.Label(frame_grafico, text="Sem dados para exibir gráficos.",
                         bg=COR_FUNDO, fg=COR_TEXTO, font=("Arial", 12))
        aviso.pack(pady=20)
        return

    # Agrupar por mês/ano
    por_mes = {}
    for d, v in zip(datas, valores):
        chave = (d.year, d.month)
        por_mes[chave] = por_mes.get(chave, 0.0) + v
    chaves_ordenadas = sorted(por_mes.keys())
    meses_labels = [f"{calendar.month_name[m]}-{a}" for (a, m) in chaves_ordenadas]
    valores_mes = [por_mes[k] for k in chaves_ordenadas]

    # Agrupar por semana ISO
    por_semana = {}
    for d, v in zip(datas, valores):
        iso = d.isocalendar()
        chave = (iso[0], iso[1])  # (anoISO, semana)
        por_semana[chave] = por_semana.get(chave, 0.0) + v
    semanas_ordenadas = sorted(por_semana.keys())
    semanas_labels = [f"{sem}ª Semana/{ano}" for (ano, sem) in semanas_ordenadas]
    valores_semana = [por_semana[k] for k in semanas_ordenadas]

    # Figura matplotlib
    fig = plt.Figure(figsize=(10, 6), dpi=100)

    # Barras por mês
    ax1 = fig.add_subplot(121)
    ax1.bar(range(len(valores_mes)), valores_mes, color=COR_DESTAQUE)
    ax1.set_title("Economia por Mês")
    ax1.set_xticks(range(len(meses_labels)))
    ax1.set_xticklabels(meses_labels, rotation=30, ha="right")
    for i, v in enumerate(valores_mes):
        ax1.text(i, v, f"R${v:.2f}", ha="center", va="bottom")

    # Pizza por semana
    ax2 = fig.add_subplot(122)
    ax2.pie(valres_semana := valores_semana, labels=semanas_labels, autopct="%1.1f%%")
    ax2.set_title("Economia por Semana")

    # Embutir no Tk
    canvas = FigureCanvasTkAgg(fig, master=frame_grafico)
    canvas.draw()
    canvas.get_tk_widget().pack(fill="both", expand=True)

# ---------- Interface ----------
janela = tk.Tk()
janela.title("APP de Poupança Pessoal")
janela.geometry("980x650")
janela.configure(bg=COR_FUNDO)

# Estilo básico do ttk (para textos claros)
style = ttk.Style()
style.theme_use("clam")
style.configure("TLabel", background=COR_FUNDO, foreground=COR_TEXTO, font=("Arial", 12))
style.configure("TButton", font=("Arial", 12))

# Frames (tk.Frame para controlar fundo preto)
frame_topo = tk.Frame(janela, bg=COR_FUNDO)
frame_topo.pack(pady=10)

frame_grafico = tk.Frame(janela, bg=COR_FUNDO)
frame_grafico.pack(fill="both", expand=True, pady=10)

# Widgets do topo
label_instrucao = tk.Label(frame_topo, text="Insira seu valor diário",
                           bg=COR_FUNDO, fg=COR_DESTAQUE, font=("Arial", 16))

# >>> ENTRY COM FUNDO PRETO <<<
entry_valor = tk.Entry(
    frame_topo,
    font=("Arial", 12),
    bg=COR_TEXTO,          # fundo preto
    fg=COR_FUNDO,          # texto branco
    insertbackground=COR_TEXTO,  # cursor branco
    relief="solid",        # borda sutil
    highlightthickness=0
)

button_salvar = ttk.Button(frame_topo, text="Salvar", command=salvar_valor)
label_status = tk.Label(frame_topo, text="", bg=COR_FUNDO, fg=COR_DESTAQUE,  font=("Arial", 11, "bold"))
label_total = tk.Label(frame_topo, text="", bg=COR_FUNDO, fg=COR_TEXTO, font=("Arial", 14, "bold"))

# Posicionamento
label_instrucao.pack()
entry_valor.pack(pady=6, ipady=3, ipadx=4)  # um pouco de padding interno
button_salvar.pack(pady=8)
label_status.pack()
label_total.pack(pady=4)

# Inicializa total e gráficos
atualizar_total()
janela.after(500, atualizar_graficos)  # carrega gráficos automaticamente

janela.mainloop()
